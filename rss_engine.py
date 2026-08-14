import os
import sys
import datetime
import requests
import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup

# Setup repository paths and GitHub Pages URL structure
REPO_ENV = os.getenv("GITHUB_REPOSITORY", "yamthouchhing/GDTRSS")
OWNER, REPO = REPO_ENV.split("/") if "/" in REPO_ENV else ("yamthouchhing", "GDTRSS")
BASE_URL = f"https://{OWNER}.github.io/{REPO}/feeds/"

EXCEL_FILE = "GDT_RSS_Feed_Gen.xlsx"
FEEDS_DIR = "feeds"
MAKE_WEBHOOK_URL = os.getenv("MAKE_WEBHOOK_URL")

os.makedirs(FEEDS_DIR, exist_ok=True)

def sanitize_filename(name):
    return "".join(c for c in name if c.isalnum() or c in ("_", "-")).strip().lower()

def create_base_rss_xml(title, link, description, items=None):
    """Generates standard RSS 2.0 XML string."""
    rss = ET.Element("rss", version="2.0")
    channel = ET.SubElement(rss, "channel")
    
    ET.SubElement(channel, "title").text = str(title)
    ET.SubElement(channel, "link").text = str(link)
    ET.SubElement(channel, "description").text = str(description)
    ET.SubElement(channel, "lastBuildDate").text = datetime.datetime.utcnow().strftime("%a, %d %b %Y %H:%M:%S GMT")
    
    if items:
        for item_data in items:
            item = ET.SubElement(channel, "item")
            ET.SubElement(item, "title").text = item_data.get("title", "No Title")
            ET.SubElement(item, "link").text = item_data.get("link", link)
            ET.SubElement(item, "description").text = item_data.get("description", "")
            ET.SubElement(item, "pubDate").text = item_data.get("pubDate", datetime.datetime.utcnow().strftime("%a, %d %b %Y %H:%M:%S GMT"))
            ET.SubElement(item, "guid").text = item_data.get("guid", item_data.get("link", link))
            
    tree = ET.ElementTree(rss)
    ET.indent(tree, space="  ")
    return tree

def generate_links():
    """Action 1: Populates RSS links into Excel sheets and creates initial feed XMLs."""
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # Process FeedLinks Sheet
    ws_feed = wb['FeedLinks']
    feed_df = pd.read_excel(EXCEL_FILE, sheet_name='FeedLinks')
    
    for idx, row in feed_df.iterrows():
        name = str(row['Name'])
        file_slug = sanitize_filename(name)
        rss_filename = f"{file_slug}.xml"
        rss_url = f"{BASE_URL}{rss_filename}"
        
        # Write RSS Link back to Sheet (Excel row index is idx + 2 due to header)
        ws_feed.cell(row=idx + 2, column=4, value=rss_url)
        
        # Generate initial XML if not exists
        filepath = os.path.join(FEEDS_DIR, rss_filename)
        if not os.path.exists(filepath):
            tree = create_base_rss_xml(title=f"{name} RSS Feed", link=row['Link'], description=f"RSS feed for {name}")
            tree.write(filepath, encoding="utf-8", xml_declaration=True)
            print(f"Created XML: {filepath}")

    # Process BundleLink Sheet
    ws_bundle = wb['BundleLink']
    bundle_df = pd.read_excel(EXCEL_FILE, sheet_name='BundleLink')
    
    bundle_url = f"{BASE_URL}bundle.xml"
    ws_bundle.cell(row=2, column=4, value=bundle_url)
    
    bundle_path = os.path.join(FEEDS_DIR, "bundle.xml")
    if not os.path.exists(bundle_path):
        b_name = bundle_df.iloc[0]['BundleName'] if not bundle_df.empty else "Bundle Feed"
        b_desc = bundle_df.iloc[0]['Description'] if not bundle_df.empty else "Aggregated Feed"
        tree = create_base_rss_xml(title=b_name, link=bundle_url, description=b_desc)
        tree.write(bundle_path, encoding="utf-8", xml_declaration=True)
        print(f"Created Bundle XML: {bundle_path}")

    wb.save(EXCEL_FILE)
    print("Excel file successfully updated with RSS URLs.")

def update_feeds():
    """Action 2: Fetches source pages, updates XML feeds, aggregates bundle, and triggers Make.com webhook."""
    print("Updating RSS feeds...")
    feed_df = pd.read_excel(EXCEL_FILE, sheet_name='FeedLinks')
    bundle_df = pd.read_excel(EXCEL_FILE, sheet_name='BundleLink')
    
    all_new_items = []
    
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    
    for _, row in feed_df.iterrows():
        name = str(row['Name'])
        source_link = str(row['Link'])
        file_slug = sanitize_filename(name)
        xml_path = os.path.join(FEEDS_DIR, f"{file_slug}.xml")
        
        items = []
        try:
            res = requests.get(source_link, headers=headers, timeout=15)
            soup = BeautifulSoup(res.text, "html.parser")
            
            # Simple content extraction (e.g., page title / social tags / posts)
            title = soup.title.string.strip() if soup.title else f"Update from {name}"
            meta_desc = soup.find("meta", property="og:description")
            desc = meta_desc["content"] if meta_desc else f"Latest content from {source_link}"
            
            items.append({
                "title": f"[{name}] {title[:80]}",
                "link": source_link,
                "description": desc,
                "pubDate": datetime.datetime.utcnow().strftime("%a, %d %b %Y %H:%M:%S GMT"),
                "guid": f"{source_link}#{datetime.datetime.utcnow().strftime('%Y%m%d%H%M')}"
            })
        except Exception as e:
            print(f"Could not fetch {source_link}: {e}")
            
        if items:
            tree = create_base_rss_xml(title=f"{name} RSS Feed", link=source_link, description=f"RSS feed for {name}", items=items)
            tree.write(xml_path, encoding="utf-8", xml_declaration=True)
            all_new_items.extend(items)

    # Update Bundle Feed
    bundle_path = os.path.join(FEEDS_DIR, "bundle.xml")
    b_name = bundle_df.iloc[0]['BundleName'] if not bundle_df.empty else "Bundle Feed"
    b_desc = bundle_df.iloc[0]['Description'] if not bundle_df.empty else "Aggregated Feed"
    
    bundle_tree = create_base_rss_xml(title=b_name, link=f"{BASE_URL}bundle.xml", description=b_desc, items=all_new_items)
    bundle_tree.write(bundle_path, encoding="utf-8", xml_declaration=True)
    print("Updated all individual feeds and bundle XML.")

    # Push Notification to Make.com AI Webhook
    if MAKE_WEBHOOK_URL:
        print("Sending push notification to Make.com...")
        payload = {
            "timestamp": datetime.datetime.utcnow().isoformat(),
            "total_items": len(all_new_items),
            "bundle_url": f"{BASE_URL}bundle.xml",
            "items": all_new_items
        }
        try:
            response = requests.post(MAKE_WEBHOOK_URL, json=payload, timeout=10)
            print(f"Make.com Webhook status: {response.status_code}")
        except Exception as e:
            print(f"Failed to trigger Make.com webhook: {e}")
    else:
        print("No MAKE_WEBHOOK_URL secret found. Skipping notification.")

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--generate":
        generate_links()
    else:
        update_feeds()
